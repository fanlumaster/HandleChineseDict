# -*- coding: utf-8 -*-
# @File  : generate_dict_02.py
# @Author: FanyFull
# @Date  : 2021/8/19

from openpyxl import load_workbook
import pymysql.cursors

# mysql connection info
connection = pymysql.connect(host='localhost',
                             user='root',
                             password='lf123456',
                             database='dictsdb',
                             cursorclass=pymysql.cursors.DictCursor)

# load xmls
workbook = load_workbook(filename='./dicts/dict_revised_2015_20210330_2.xlsx')
# print(workbook.sheetnames)
worksheet = workbook['資料表2']
# cell_value = worksheet['K5'].value
# print(cell_value)
# print(worksheet.values)

n = 0
with connection:
    with connection.cursor() as cursor:
        print('start insert')
        n = 0
        for row in worksheet.values:
            n = n + 1
            if n == 1:
                continue
            # if n > 10:
            #     break
            # insert records
            sql = "INSERT INTO `dict_tb2` (`property`, `word_num`, `name`, `part_excluded`, `part_num`, `all_num`, `punc`, `pinyin`, `similar`, `opposite`, `meaning`, `edit_comment`, `duoyin`, `yiti`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            cursor.execute(sql, row)
            print('第 ' + str(n - 1) + ' 条数据插入完毕')
        print('info prepared ends')
    # print(row)

    connection.commit()
