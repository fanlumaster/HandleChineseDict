# -*- coding: utf-8 -*-
# @File  : generate_dict.py
# @Author: FanLu
# @Date  : 2021/8/20

from openpyxl import load_workbook
from openpyxl import *

workbook = load_workbook(filename='./dicts/dict_revised_2015_20210330_1.xlsx')
# print(workbook.sheetnames)
worksheet = workbook['工作表1']
# cell_value = worksheet['K5'].value
# print(cell_value)
print(worksheet.values)
n = 0
for row in worksheet.values:
    for value in row:
        print(value)
    print('\n==========')
    print('==========')
    n = n + 1
    if n > 100:
        break