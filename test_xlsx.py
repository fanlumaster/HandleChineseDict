# -*- coding: utf-8 -*-
# @File  : test_xlsx.py
# @Author: FanyFull
# @Date  : 2021/8/20

# 测试 openpyxl 的使用

from openpyxl import load_workbook

# load xmls
workbook = load_workbook(filename='./dicts/dict_revised_2015_20210330_1.xlsx')
# print(workbook.sheetnames)
worksheet = workbook['工作表1']
cell_value = worksheet['K5'].value
# print(cell_value)
# print(worksheet.values)

n = 0
max_len = 20
column_num = 2
for row in worksheet.values:
    n = n + 1
    if len(str(row[column_num])) > 20:
        if len(str(row[column_num])) > max_len:
            max_len = len(str(row[column_num]))
        print(row[column_num])
        print('========')

print(max_len)